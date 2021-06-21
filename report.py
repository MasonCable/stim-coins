import openpyxl

headers = ['id', 'currency', 'balance', 'hold', 'available', 'profile_id', 'trading_enabled']
myWb = openpyxl.Workbook()
mySheet = myWb.active
accounts = client.accounts()
columnN = 1

for i in range(len(headers)):
    cell = mySheet.cell(row=1, column=columnN)
    cell.value = headers[i]
    columnN += 1
myWb.save('accounts.xlsx')

newWb = openpyxl.load_workbook('accounts.xlsx')
newSheet = newWb.active
print(newSheet)